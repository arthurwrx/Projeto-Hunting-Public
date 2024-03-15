import time
import os
from openpyxl import load_workbook
import openpyxl
import pyautogui
from selenium.webdriver.firefox.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By 
import PySimpleGUI as sg

service = Service(executable_path="./geckodriver.exe")
options = webdriver.FirefoxOptions()
driver = webdriver.Firefox(service=service, options=options)

def email_senha():
    global email
    email = pyautogui.prompt(text='Digite o seu email do Linkedin', title='' , default='')
    global senha 
    senha = pyautogui.prompt(text='Digite sua senha', title='' , default='')

email_senha()

### Listas 
lista_tipos_empregos = ["Tempo integral","Meio período","Autônomo","Freelance","Temporário","Estágio","Aprendiz", "Terceirizado"]

def start():

    global nome_planilha

    if __name__ == "__main__":
        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)

        layout = [
            # [sg.Column([[sg.Image(r'logo\logo-assinatura.png')]], justification='center')],
            [sg.Column([[sg.Text('Bem vindo a Automação de Hunting Linkedin.',font=('Helvetica', 12 ,'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Antes de começar, insira os perfis dos linkedins que você quer huntiar na planilha "Lista de Perfis"',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Salve a planilha com os perfis, depois feche-a, para não dar erro na automação.',font=('Helvetica', 10, ))]], justification='center')],
            [sg.Column([[sg.Text('Assim que o programa finalizar todas as buscas, os resultados estarão na planilha com o nome que você escolheu.',font=('Helvetica', 10, ))]], justification='center')],

            [sg.Text('Nome da planilha desejada '), sg.InputText(key="nome_planilha")],
            [
            sg.Column([[sg.Button('Começar!', size=tamanho_botao, font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center')
        ]]

        window = sg.Window('Linkedin').layout(layout)

        while True: 
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                break

            elif event == 'Começar!':
                window.close()
                nome_planilha = values['nome_planilha']
                cria_excel()
                login()
                body()
                mensagem_final()


def login():

    time.sleep(2)
    driver.get('https://www.linkedin.com/login')
    driver.find_element(By.XPATH,'//*[@id="username"]').send_keys(email)
    driver.find_element(By.XPATH,'//*[@id="password"]').send_keys(senha)
    time.sleep(1)
    driver.find_element(By.XPATH,"//*[@type='submit']").click()
    link_atual = ""
    while link_atual != "https://www.linkedin.com/feed/":
        link_atual = driver.current_url

def body():

    nomeCaminhoArquivo = r'Lista de Perfis.xlsx'
    planilha_aberta =  load_workbook(filename= nomeCaminhoArquivo)
    sheet_selecionada = planilha_aberta["Plan1"] ## Acessa a ab a que estiver dentro de wb

    for linha in range(2, len(sheet_selecionada['A']) + 1):

        global profile_url

        profile_url = sheet_selecionada[f'A{linha}'].value

        if profile_url == None:
            break

        if profile_url[-1] != "/":
            profile_url = profile_url + "/"

        economiza_memoria()

        time.sleep(5)

        try:
            pagina_nao_existe = driver.find_element_by_xpath("//h2[contains(@class, 'artdeco-empty-state__headline')]")
            pagina_nao_existe = pagina_nao_existe.text

            if pagina_nao_existe == "Esta página não existe":
                toExcel_casonaoexista()
                continue
        
        except:
             pass
            
        try:
            pagina_nao_encontrada = driver.find_element_by_xpath("//h1[@class='heading' and contains(text(), 'Page not found')]")
            pagina_nao_encontrada = pagina_nao_encontrada.text

            if pagina_nao_encontrada == "Page not found":
                toExcel_casonaoexista()
                continue
        except:
            pass

        try:
            pagina_nao_encontrada = driver.find_element_by_xpath("//h1[@class='heading' and contains(text(), 'Página não encontrada')]")
            pagina_nao_encontrada = pagina_nao_encontrada.text

            if pagina_nao_encontrada == "Página não encontrada":
                toExcel_casonaoexista()
                continue
        except:
            pass

        ## Verifica se o link atual existe.
        if link_atual == "https://www.linkedin.com/404/":
            toExcel_casonaoexista()
            continue

        else:
            getPerfil()
            captura_cargos()
            getEducation()
            getLanguages()
            toExcel()

def economiza_memoria():

    global link_atual

    driver.get("https://www.google.com")

    # Abre uma nova aba usando JavaScript
    driver.execute_script("window.open('', '_blank');")

    # Aguarda até que o número de janelas seja igual a 2
    while len(driver.window_handles) != 2:
        time.sleep(1)

    # Obtém a lista de identificadores de janelas (handles)
    janelas = driver.window_handles

    # Alterna para a nova aba
    driver.switch_to.window(janelas[1])

    # Abre uma página na nova aba
    driver.get(profile_url)

    # Espera algum tempo (opcional)
    # Fecha a aba anterior
    driver.switch_to.window(janelas[0])
    driver.close()
    driver.switch_to.window(janelas[1])

    link_atual = driver.current_url
    print(link_atual)
    ##continue

def getPerfil():

    global nome_candidato 
    global localizacao_candidato
    global atuacao_candidato

    try:
        nome_candidato = str(driver.find_element(By.XPATH,"//*[@class='text-heading-xlarge inline t-24 v-align-middle break-words']").text).strip()
    except:
        nome_candidato = "Não foi possível capturar o nome do candidato"
    try:
        localizacao_candidato = str(driver.find_element(By.XPATH,"//*[@class='text-body-small inline t-black--light break-words']").text).strip()
    except:
        localizacao_candidato = "Não foi possível capturar o local do candidato"
    try:
        atuacao_candidato = str(driver.find_element(By.XPATH,"//*[@class='text-body-medium break-words']").text).strip()
    except:
        atuacao_candidato = "Não foi possível capturar a atuação do candidato"

def captura_cargos():

    global nome_empresa1
    global nome_empresa2
    global nome_empresa3

    global nome_cargo1_empresa1
    global nome_cargo2_empresa1
    global nome_cargo3_empresa1
    global nome_cargo1_empresa2
    global nome_cargo2_empresa2
    global nome_cargo3_empresa2
    global nome_cargo1_empresa3
    global nome_cargo2_empresa3
    global nome_cargo3_empresa3    

    global periodo_empresa1_cargo1
    global periodo_empresa1_cargo2
    global periodo_empresa1_cargo3

    global periodo_empresa2_cargo1
    global periodo_empresa2_cargo2
    global periodo_empresa2_cargo3

    global periodo_empresa3_cargo1
    global periodo_empresa3_cargo2
    global periodo_empresa3_cargo3  

    time.sleep(5)

    nomes_empresas = []
    nomes_cargos = []
    periodos = []
    
    driver.get(profile_url+"details/experience/")
    time.sleep(4)

    nomes_empresas_com_promocao_selector = driver.find_elements(By.XPATH,'//li[@class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')

    for i in range(3): ## O For só vai rodar 3 vezes, porque só importa até 3 empresas

        try:

            nome_empresa = nomes_empresas_com_promocao_selector[i].text
            partes_nome = nome_empresa.rsplit(maxsplit=1)

            dados_separados = [item.split('\n') for item in partes_nome]

            comprimento_sublista = len(dados_separados[0])


            nova_lista = []

            for i in range(1,comprimento_sublista,2):

                nova_lista.append(dados_separados[0][i])
        
        except:
            break

        nome_empresa_sem_promocao = nova_lista[2]

        # Para casos de empresa sem promoção ## FUNCIONANDO PERFEITAMENTE NÃO MEXER

        if "·" in nome_empresa_sem_promocao:

            nome_empresa_sem_promocao = nova_lista[1]
            nome_empresa_sem_promocao = nome_empresa_sem_promocao.split("·")
            nome_empresa_sem_promocao = nome_empresa_sem_promocao[0].strip()

            nomes_empresas.append(nome_empresa_sem_promocao)

            cargo = nova_lista[0]

            nomes_cargos.append(cargo)

            tempo_permanencia = nova_lista[2]
            tempo_permanencia = tempo_permanencia.split("·")
            tempo_permanencia = tempo_permanencia[0].strip()

            periodos.append(tempo_permanencia)

            nomes_cargos.append("Não tem")
            nomes_cargos.append("Não tem")
            periodos.append("Não tem")
            periodos.append("Não tem")


            # print(f"Então essa é uma empresa que não teve promoção e o seu nome é: {nome_empresa_sem_promocao} e seu cargo mais recente é: {nova_lista[0]}, ficou no período de {tempo_permanencia}")

        ### Para casos de empresas com promoção
        
        else:

            if "·" in nova_lista[3]: ## Aparentemente está funcionando corretamente, o problema é quando a pessoa coloca descrição nos cargos.
                
                nome_empresa_com_promocao = nova_lista[0]
                nomes_empresas.append(nome_empresa_com_promocao)

                cargo = nova_lista[2]
                nomes_cargos.append(cargo)

                periodo = nova_lista[3]
                periodos.append(periodo)


                # print(f"Então essa é uma empresa que teve promoção e o seu nome é: {nome_empresa_com_promocao} e seu cargo mais recente é: {cargo}, trabalhou no período de {periodo}")
                
                i = 4
                # print(nova_lista)

                ## Segundo Cargo

                try:
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1

                    ponto_parada = i
                    posicao_periodo_segundo_cargo = i

                    segundo_cargo = nova_lista[i-1]

                    if segundo_cargo in lista_tipos_empregos:
                        segundo_cargo = nova_lista[i-2]
                        nomes_cargos.append(segundo_cargo)
                    
                    else:
                        segundo_cargo = nova_lista[i-1]
                        nomes_cargos.append(segundo_cargo)

                    tempo_permanencia = nova_lista[i]
                    periodos.append(tempo_permanencia)
                    
                
                except:
                    nomes_cargos.append("Não foi possível capturar informações")
                    periodos.append("Não foi possível capturar informações")

                ## TERCEIRO CARGO

                try:
                    
                    i = ponto_parada
                    i += 1
    
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1
                    
                    # if i - 1 == posicao_periodo_segundo_cargo:
                    #     nomes_cargos.append("Não tem")
                    #     periodos.append("Não tem")
                    #     pass
                    
                    else:
                        posicao_terceiro_cargo = i 
                            
                        terceiro_cargo = nova_lista[i-1]
                        
                        nomes_cargos.append(terceiro_cargo)
                    
                        periodo.append(posicao_terceiro_cargo)

                        print(f"O Terceiro cargo dele está na posição {posicao_terceiro_cargo} e o cargo é {terceiro_cargo}")

                except:
                    nomes_cargos.append("Não tem")
                    periodos.append("Não tem")





            elif "·" in nova_lista[4]:  ### CASOS DE EMPRESA COM PROMOÇÕES
                ### Primeiro Cargo

                nome_empresa_com_promocao = nova_lista[0]
                nomes_empresas.append(nome_empresa_com_promocao)

                primeiro_cargo = nova_lista[3]

                if primeiro_cargo in lista_tipos_empregos:
                    primeiro_cargo = nova_lista[2]
                    nomes_cargos.append(primeiro_cargo)
                
                else:
                    primeiro_cargo = nova_lista[3]
                    nomes_cargos.append(primeiro_cargo)

                periodo = nova_lista[4]
                periodos.append(periodo)
                
                i = 5

                ## Segundo cargo
                try:
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1

                    posicao_periodo_segundo_cargo = i

                    segundo_cargo = nova_lista[i-1]

                    if segundo_cargo in lista_tipos_empregos:
                        segundo_cargo = nova_lista[i-2]
                        nomes_cargos.append(segundo_cargo)
                    
                    else:
                        segundo_cargo = nova_lista[i-1]
                        nomes_cargos.append(segundo_cargo)

                    tempo_permanencia = nova_lista[i]
                    periodos.append(tempo_permanencia)
                    
                
                except:
                    nomes_cargos.append("Não foi possível capturar informações")
                    periodos.append("Não foi possível capturar informações")

                ## TERCEIRO CARGO

                try:
                    i = ponto_parada
                    i += 1
    
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1
                    
                    if i - 1 == posicao_periodo_segundo_cargo:
                        nomes_cargos.append("Não tem")
                        periodos.append("Não tem")
                        pass
                    
                    else:
                        posicao_terceiro_cargo = i 
                            
                        terceiro_cargo = nova_lista[i-1]
                        
                        nomes_cargos.append(terceiro_cargo)
                    
                        periodo.append(posicao_terceiro_cargo)

                        # print(f"O Terceiro cargo dele está na posição {posicao_terceiro_cargo} e o cargo é {terceiro_cargo}")

                except:

                    nomes_cargos.append("Não tem")
                    periodos.append("Não tem")


                
            elif nova_lista[4] in lista_tipos_empregos: ## CASO DE O PONTO FOR INTEGRAL E ETC
                
                nome_empresa_com_promocao = nova_lista[0]
                nomes_empresas.append(nome_empresa_com_promocao)

                cargo = nova_lista[3]
                nomes_cargos.append(cargo)

                periodo = nova_lista[5]
                periodos.append(periodo)
                
                i = 6

                ## Segundo Cargo

                try:
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1

                    posicao_periodo_segundo_cargo = i

                    segundo_cargo = nova_lista[i-1]

                    if segundo_cargo in lista_tipos_empregos:
                        segundo_cargo = nova_lista[i-2]
                        nomes_cargos.append(segundo_cargo)
                    
                    else:
                        segundo_cargo = nova_lista[i-1]
                        nomes_cargos.append(segundo_cargo)

                    tempo_permanencia = nova_lista[i]
                    periodos.append(tempo_permanencia)
                    
                
                except:
                    nomes_cargos.append("Não foi possível capturar informações")
                    periodos.append("Não foi possível capturar informações")

                ## TERCEIRO CARGO

                try:
                    
                    i = ponto_parada
                    i += 1
    
                    while not ("meses" in nova_lista[i] or "ano" in nova_lista[i] or "anos" in nova_lista[i] or "mês" in nova_lista[i]):
                        i += 1
                    
                    if i - 1 == posicao_periodo_segundo_cargo:
                        nomes_cargos.append("Não tem")
                        periodos.append("Não tem")
                        pass
                    
                    else:
                        posicao_terceiro_cargo = i 
                            
                        terceiro_cargo = nova_lista[i-1]
                        
                        nomes_cargos.append(terceiro_cargo)
                    
                        periodo.append(posicao_terceiro_cargo)

                        # print(f"O Terceiro cargo dele está na posição {posicao_terceiro_cargo} e o cargo é {terceiro_cargo}")

                except:
                    nomes_cargos.append("Não tem")
                    periodos.append("Não tem")

                

    # print(nomes_empresas)
    # print(nomes_cargos)
    # print(periodos)
    # print()
    # print()
    # print(len(nomes_empresas))
    # print(len(nomes_cargos))
    # print(len(periodos))

    # print("----------------------")

    nome_empresa1 = nomes_empresas[0]

    nome_cargo1_empresa1 = nomes_cargos[0]
    nome_cargo2_empresa1 = nomes_cargos[1]
    nome_cargo3_empresa1 = nomes_cargos[2]
    
    periodo_empresa1_cargo1 = periodos[0]
    periodo_empresa1_cargo2 = periodos[1]
    periodo_empresa1_cargo3 = periodos[2]

    try:
        nome_empresa2 = nomes_empresas[1]

        nome_cargo1_empresa2 = nomes_cargos[3]
        nome_cargo2_empresa2 = nomes_cargos[4]
        nome_cargo3_empresa2 = nomes_cargos[5]

        periodo_empresa2_cargo1 = periodos[3]
        periodo_empresa2_cargo2 = periodos[4]
        periodo_empresa2_cargo3 = periodos[5]
    
    except:

        nome_empresa2 = "Não tem"

        nome_cargo1_empresa2 = "Não tem"
        nome_cargo2_empresa2 = "Não tem"
        nome_cargo3_empresa2 = "Não tem"

        periodo_empresa2_cargo1 = "Não tem"
        periodo_empresa2_cargo2 = "Não tem"
        periodo_empresa2_cargo3 = "Não tem"

    try:

        nome_empresa3 = nomes_empresas[2]

        nome_cargo1_empresa3 = nomes_cargos[6]
        nome_cargo2_empresa3 = nomes_cargos[7]
        nome_cargo3_empresa3 = nomes_cargos[8]

        periodo_empresa3_cargo1 = periodos[6]
        periodo_empresa3_cargo2 = periodos[7]
        periodo_empresa3_cargo3 = periodos[8]
    
    except:

        nome_empresa3 = "Não tem"

        nome_cargo1_empresa3 = "Não tem"
        nome_cargo2_empresa3 = "Não tem"
        nome_cargo3_empresa3 = "Não tem"

        periodo_empresa3_cargo1 = "Não tem"
        periodo_empresa3_cargo2 = "Não tem"
        periodo_empresa3_cargo3 = "Não tem"

def getEducation():

    global nome_faculdade1
    global nome_faculdade2
    global nome_faculdade3
    global nome_curso1
    global nome_curso2
    global nome_curso3
    global periodo_curso1
    global periodo_curso2
    global periodo_curso3

    nomes_faculdades = []
    nomes_cursos = []
    tempos_formacoes = []


    driver.get(profile_url+"details/education/")
    time.sleep(3)

    

    formacoes_selector = driver.find_elements(By.XPATH,'//li[@class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')

    ## Esse for foi pensado para pegar as 3 últimas formações.

    for i in range(3): 

        try:

            nome_formacao = formacoes_selector[i].text
            partes_nome = nome_formacao.rsplit(maxsplit=1)

            dados_separados = [item.split('\n') for item in partes_nome]

            comprimento_sublista = len(dados_separados[0])
            # print("O tamanho da sublista é: " + str(comprimento_sublista))  

            nova_lista = []

            for i in range(1,comprimento_sublista,2):

                nova_lista.append(dados_separados[0][i])

        except:
            break

        nome_faculdade = nova_lista[0]
        nomes_faculdades.append(nome_faculdade)

        try:
            nome_curso = nova_lista[1]
            nomes_cursos.append(nome_curso)

        except:
            nome_curso = "Não colocou o nome do curso"
            nomes_cursos.append(nome_curso)

        try:
            tempo_formacao = nova_lista[2]
            tempos_formacoes.append(tempo_formacao)
        except:
            tempo_formacao = "Não colocou o tempo de formação"
            tempos_formacoes.append(tempo_formacao)

###########################################################
        
    try:    
        nome_faculdade1 = nomes_faculdades[0]
        nome_curso1 = nomes_cursos[0]
        periodo_curso1 = tempos_formacoes[0]
    
    except:
        nome_faculdade2 = "Não tem"
        nome_curso2 = "Não tem"
        periodo_curso2 = "Não tem"

    try:
        nome_faculdade2 = nomes_faculdades[1]
        nome_curso2 = nomes_cursos[1]
        periodo_curso2 = tempos_formacoes[1]
    
    except IndexError:
        nome_faculdade2 = "Não tem"
        nome_curso2 = "Não tem"
        periodo_curso2 = "Não tem"

    try:
        nome_faculdade3 = nomes_faculdades[2]
        nome_curso3 = nomes_cursos[2]
        periodo_curso3 = tempos_formacoes[2]
    
    except IndexError:
        nome_faculdade3 = "Não tem"
        nome_curso3 = "Não tem"
        periodo_curso3 = "Não tem"

    # print(nome_faculdade1)
    # print(nome_curso1)
    # print(periodo_curso1)

    # print(nome_faculdade2)
    # print(nome_curso2)
    # print(periodo_curso2)
    
    # print(nome_faculdade3)
    # print(nome_curso3)
    # print(periodo_curso3)
         
def getLanguages():

    global idioma1
    global idioma2
    global idioma3 
    global nivel_idioma1
    global nivel_idioma2
    global nivel_idioma3

    driver.get(profile_url+"details/languages/")
    time.sleep(2)

    nomes_idiomas_selector = driver.find_elements(By.XPATH,"//div[contains(concat(' ', normalize-space(@class), ' '), ' display-flex align-items-center mr1 t-bold ')]/span[1]")
    niveis_idiomas_selector = driver.find_elements(By.XPATH,"//*[@class='t-14 t-normal t-black--light']/span[1]")
    nao_existe_idiomas = driver.find_element(By.XPATH,'//li[@class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')

    nada_por_enquanto = nao_existe_idiomas.text

    if "Nada para ver por enquanto" in nada_por_enquanto:

        idioma1 = "Não tem"
        nivel_idioma1 = "Não tem"
        idioma2 = "Não tem"
        nivel_idioma2 = "Não tem"
        idioma3 = "Não tem"
        nivel_idioma3 = "Não tem"
        

    else:

        try:
            nome_idioma1 = nomes_idiomas_selector[0]
            idioma1 = nome_idioma1.text
            nivel_idioma1 = niveis_idiomas_selector[0]
            nivel_idioma1 = nivel_idioma1.text

        except:
            nome_idioma1 = "Não tem"
            nivel_idioma1 = "Não tem"

        try: ## Bloco do segundo idioma se houver

            nome_idioma2 = nomes_idiomas_selector[1]
            idioma2 = nome_idioma2.text

            nivel_idioma2 = niveis_idiomas_selector[1]
            nivel_idioma2 = nivel_idioma2.text

        except:
            idioma2 = "Não tem"
            nivel_idioma2 = "Não tem"

        try:
            nome_idioma3 = nomes_idiomas_selector[2]
            idioma3 = nome_idioma3.text

            nivel_idioma3 = niveis_idiomas_selector[2]
            nivel_idioma3 = nivel_idioma3.text

        except:
            idioma3 = "Não tem"
            nivel_idioma3 = "Não tem"

def cria_excel():

    global planilha

    workbook = openpyxl.Workbook()

    workbook.create_sheet(title="Hunting")

    if "Sheet" in workbook.sheetnames:
        workbook.remove_sheet(workbook["Sheet"])

    workbook.save(filename=f"{nome_planilha}.xlsx")

    planilha = f'{nome_planilha}.xlsx'

def toExcel():

    # print("passei por aqui")

    arquivo = planilha

    wb = load_workbook(arquivo)
    
    ws = wb.active

    ws['A1'].value = 'Linkedin'
    ws['B1'].value = 'Nome'
    ws['C1'].value = 'Local'
    ws['D1'].value = 'Atuação'
    ws['E1'].value = 'Nome Faculdade 1'
    ws['F1'].value = 'Curso Faculdade 1'
    ws['G1'].value = 'Período Faculdade 1'
    ws['H1'].value = 'Nome Faculdade 2'
    ws['I1'].value = 'Curso Faculdade 2'
    ws['J1'].value = 'Período Faculdade 2'
    ws['K1'].value = 'Nome Faculdade 3'
    ws['L1'].value = 'Curso Faculdade 3'
    ws['M1'].value = 'Período Faculdade 3'

    ws['N1'].value = 'Empresa 1'

    ws['O1'].value = 'Cargo 1 Empresa 1'
    ws['P1'].value = 'Período Cargo 1 Empresa 1'

    ws['Q1'].value = 'Cargo 2 Empresa 1'
    ws['R1'].value = 'Período Cargo 2 Empresa 1'

    ws['S1'].value = 'Cargo 3 Empresa 1'
    ws['T1'].value = 'Período Cargo 3 Empresa 1'

    ws['U1'].value = 'Empresa 2'

    ws['V1'].value = 'Cargo 1 Empresa 2'
    ws['W1'].value = 'Período Cargo 1 Empresa 2'

    ws['X1'].value = 'Cargo 2 Empresa 2'
    ws['Y1'].value = 'Período Cargo 2 Empresa 2'

    ws['Z1'].value = 'Cargo 3 Empresa 2'
    ws['AA1'].value = 'Período Cargo 3 Empresa 2'

    ws['AB1'].value = 'Empresa 3'

    ws['AC1'].value = 'Cargo 1 Empresa 3'
    ws['AD1'].value = 'Período Cargo 1 Empresa 3'

    ws['AE1'].value = 'Cargo 2 Empresa 3'
    ws['AF1'].value = 'Período Cargo 2 Empresa 3'

    ws['AG1'].value = 'Cargo 3 Empresa 3'
    ws['AH1'].value = 'Período Cargo 3 Empresa 3'

    ws['AI1'].value = 'Idioma1'
    ws['AJ1'].value = 'Nivel1'

    ws['AK1'].value = 'Idioma2'
    ws['AL1'].value = 'Nivel2'

    ws['AM1'].value = 'Idioma3'
    ws['AN1'].value = 'Nivel3'

    ws.append({'A':profile_url, 
    'B':nome_candidato, 
    'C':localizacao_candidato, 
    'D':atuacao_candidato, 
    'E':nome_faculdade1, 
    'F':nome_curso1, 
    'G':periodo_curso1, 
    'H':nome_faculdade2, 
    'I':nome_curso2, 
    'J':periodo_curso2, 
    'K':nome_faculdade3, 
    'L':nome_curso3, 
    'M':periodo_curso3, 

    'N':nome_empresa1, 

    'O':nome_cargo1_empresa1, 
    'P':periodo_empresa1_cargo1, 

    'Q':nome_cargo2_empresa1, 
    'R':periodo_empresa1_cargo2, 

    'S':nome_cargo3_empresa1,
    'T':periodo_empresa1_cargo3,

    'U':nome_empresa2, 

    'V':nome_cargo1_empresa2, 
    'W':periodo_empresa2_cargo1, 

    'X':nome_cargo2_empresa2, 
    'Y':periodo_empresa2_cargo2,

    'Z':nome_cargo3_empresa2,
    'AA':periodo_empresa2_cargo3,

    'AB':nome_empresa3,

    'AC':nome_cargo1_empresa3, 
    'AD':periodo_empresa3_cargo1,

    'AE':nome_cargo2_empresa3,
    'AF':periodo_empresa3_cargo2,

    'AG':nome_cargo3_empresa3,
    'AH':periodo_empresa3_cargo3,

    'AI':idioma1,
    'AJ':nivel_idioma1,

    'AK':idioma2,
    'AL':nivel_idioma2,

    'AM':idioma3,
    'AN':nivel_idioma3

    })

    wb.save(arquivo)

def toExcel_casonaoexista():
     
    arquivo = planilha

    wb = load_workbook(arquivo)
    
    ws = wb.active

    ws.append({'A':profile_url,
                
    'B':"O link do perfil desse candidato não existe",
    'C':"O link do perfil desse candidato não existe",
    'D':"O link do perfil desse candidato não existe",
    'E':"O link do perfil desse candidato não existe",
    'F':"O link do perfil desse candidato não existe",
    'G':"O link do perfil desse candidato não existe",
    'H':"O link do perfil desse candidato não existe",
    'I':"O link do perfil desse candidato não existe",
    'J':"O link do perfil desse candidato não existe",
    'K':"O link do perfil desse candidato não existe",
    'L':"O link do perfil desse candidato não existe",
    'M':"O link do perfil desse candidato não existe",

    'N': "O link do perfil desse candidato não existe",

    'O':"O link do perfil desse candidato não existe",
    'P':"O link do perfil desse candidato não existe",

    'Q':"O link do perfil desse candidato não existe",
    'R':"O link do perfil desse candidato não existe",

    'S':"O link do perfil desse candidato não existe",
    'T':"O link do perfil desse candidato não existe",

    'U':"O link do perfil desse candidato não existe",

    'V':"O link do perfil desse candidato não existe",
    'W':"O link do perfil desse candidato não existe",

    'X':"O link do perfil desse candidato não existe",
    'Y':"O link do perfil desse candidato não existe",

    'Z':"O link do perfil desse candidato não existe",
    'AA':"O link do perfil desse candidato não existe",

    'AB':"O link do perfil desse candidato não existe",

    'AC':"O link do perfil desse candidato não existe",
    'AD':"O link do perfil desse candidato não existe",

    'AE':"O link do perfil desse candidato não existe",
    'AF':"O link do perfil desse candidato não existe",

    'AG':"O link do perfil desse candidato não existe",
    'AH':"O link do perfil desse candidato não existe",

    'AI':"O link do perfil desse candidato não existe",
    'AJ':"O link do perfil desse candidato não existe",

    'AK':"O link do perfil desse candidato não existe",
    'AL':"O link do perfil desse candidato não existe",

    'AM':"O link do perfil desse candidato não existe",
    'AN':"O link do perfil desse candidato não existe"
    })

    wb.save(arquivo)

def mensagem_final():

    global nome_planilha

    if __name__ == "__main__":
        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)

        layout = [
            [sg.Column([[sg.Image(r'logo\logo-assinatura.png')]], justification='center')],

            [sg.Column([[sg.Text('Processo Finalizado!',font=('Calibri', 14 ,'bold'))]], justification='center')],
            [sg.Column([[sg.Text(f'Confira a planilha {nome_planilha} na pasta do programa.',font=('Helvetica', 12, 'bold'))]], justification='center')],
            [sg.Column([[sg.Button('Ok', size=tamanho_botao, font=('Calibri', 12, 'bold'))]], justification='center', element_justification='center')
        ]]

        window = sg.Window('Linkedin').layout(layout)

        while True: 
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                break

            elif event == 'Ok':
                window.close()
                driver.quit()

start()




